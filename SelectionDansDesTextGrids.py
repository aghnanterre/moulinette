#!/usr/bin/python 
# -*- coding: utf-8 -*-
#Filename: 2018_06_06_SelectionDansDesTextGrids2.py

#iso-8859-1

# programme fait par Amal Guha, Modyco (CNRS UMR7114 - Université Paris Nanterre)

#####################################################################
#####################################################################
#
#	Le programme se lance avec une requête en argument (ce sera argv[1]); La fonction expansion définit la liste des requetes correspondant à la argv[1] (flexions).
# Si cette liste n'est pas explicitement prévue par la fonction expansion, la fonction expansion renvoie une liste à un élément qui contient argv[1].
#
#  Ce programme crée un fichier csv (dans le dossier resultats) qui reprend tous les tours de parole extraits d'un textgrid,
#  qui contiennent la chaîne de caractères mise dans la variable laRequete. Il travaille à partir du répertoire corpus où sont tous les fichiers .Textgrid (repertoireFichierTraite).  
# pour vérification que les Textrids ont bien été lus par le programme, on me dans le répertoire listesFichiersTraites/Date un csv qui donne le nom des transcriptions traitées et le nombre de mots traités pour chacune.
#
#####################################################################

#################################################################################
#fonction def SR()
# où sont définies exhaustivement les sous-requêtes pertinentes
#r#################################################################################

def SR():
	sousRequetes=list()
#	sousRequetes=["à nous","A nous","de nous","De nous","pour nous","Pour nous","Chez nous","chez nous","Avec nous","avec nous","sans nous","Sans nous","parmi nous","Parmi nous","Entre nous","entre nous","Sur nous","sur nous"]
	#sousRequetes=["Tu nous","tu nous","Il nous","il nous","Elle nous","elle nous","On nous","on nous","Nous nous","nous nous","Vous nous","vous nous","Ils nous","ils nous","Elles nous","elles nous"]
	sousRequetes=["Comme nous","comme nous","comme vous","Comme vous"]
	return sousRequetes

##########################################################################################################################################
# FONCTION PRINCIPALE : qui imprime les nom d'enquête et tour(s) de parole où figure la string maRequete ##########################################################################################################################################

#################################################################################
#fonction retourRequete(repertoireFichierTraite:string, maRequete : string, nomEnquete : string, 
#nomRepertoireFichiersResultats : string
#retourne une liste de trois valeurs : [resuCSV, nomEnquete, cptLignes] : cptLignes permet de vérifier que le fichier a bien été lu par le programme
#################################################################################

def retourRequete(repertoireFichierTraite,maRequete,nomEnquete,nomRepertoireFichiersResultats,fichierLog):
	# Ouverture du fichier sur lequel effectuer la requête
	# fichier Texte exporté depuis Praat au format txt, codage utf-8


#	print "*** Début de la phase d'ouverture" # Debug
	
#	print "Essai d'ouverture du fichier dont le nom est fourni en quatrième paramètre: ",nomEnquete # Debug
	try :
		fo = open(repertoireFichierTraite+"/"+nomEnquete, "r")
	except IOError:
		fichierLog.write("Le fichier dont le nom est le quatrième paramètre ("+repertoireFichierTraite+"/"+nomEnquete+") ne s'ouvre pas !")

	# Initialisation des variables
	nomTrouve=0
	nbOcc=0

	#---------------------------------------------------		
	#BOUCLE PRINCIPALE (de traitement du fichier Textgrid entree.txt)
	#il faut remettre le pointeur au début du fichier

	fo.seek(0)
	#print(maRequete," cherché dans l'enquête " ,nomEnquete)
	indiceNomCourant=-1
	cptLignes=0
	resuCSV=""
	nbMotsEnquete=0
	for line in fo:
		cptLignes=cptLignes+1
	# Identification du locuteur (nomTire)
		trouveNomEnPosition = line.find("name =")
		if trouveNomEnPosition>0:
			nbMotsTourDeParole=0
			nomTrouve=1
			indiceNomCourant +=1
			line= line[trouveNomEnPosition+8:]
			nomTire=line[0:len(line)-3]

	# positionIntervalle (coordonnée x du début de l'intervalle)		
		trouveXEnPosition= line.find("xmin")
		if trouveXEnPosition>0:
			line= line[trouveNomEnPosition+12:]
			xDeb=line[8:15]
	
	# lorsque la ligne contient "text"		
		trouveTexteEnPosition = line.find("text")
		
		if trouveTexteEnPosition>0:
			line2= line[trouveTexteEnPosition+8:]
			texte=line2[0:len(line2)-3]
	# Si cette ligne du Textgrid correspond à un intervalle (si texte !="")
			# C'est ici qu'on regarde si maRequete est dans le tour de parole
			if (texte!=""):
				laSousRequete=""
			# suppression des signes de chevauchement qui risquent de couper les mots ou expressions de maRequete
				texte=texte.replace('<','')
				texte=texte.replace('>','')
			#find proprement dit
				indiceTourDeParole=-len(maRequete) # indice du tour de parole entier (text) à partir duquel commence la variable reste 
#				nbMotsTourDeParole=len(reste.split())	
				if texte.find(maRequete)>0:
					reste=texte.lower()
					resteCalcule=""
					indiceDeReste=0 # indice dans le tour de parole entier (text) à partir duquel commence la variable reste
					
				# chetcher par find())	
					prochaineOccurrenceDansReste=reste.find(maRequete)
					while prochaineOccurrenceDansReste>=0:
					#a exclure ?
						aExclure=''
						indiceTourDeParole=indiceTourDeParole+len(maRequete)+prochaineOccurrenceDansReste
				#		aExclure=texte[indiceTourDeParole:indiceTourDeParole+len(maRequete)]# donne exactement maRequete
						if texte[indiceTourDeParole:indiceTourDeParole+len(maRequete)+10].find("n\'importe")>0:
							aExclure="n\'importe"
						if texte[indiceTourDeParole:indiceTourDeParole+len(maRequete)+14].find("ne serait-ce")>0:
							aExclure="ne serait-ce"
						if texte[indiceTourDeParole:indiceTourDeParole+len(maRequete)+8].find("n\'dour")>0:
							aExclure="n\'dour"
														
							
						indexRE=-1
					# chercher par Regex (inclus dans la boucle du find)
						stringRE = "([^ ]+\s+)*suis (([^ ]+(\s+)){0,3}|.)" #"([^ ]+\s+)*comme|Comme ([^ ]+\s+){0,1}.|( ?)$" # stringRE 
						p=re.compile(stringRE)
						m= p.match(reste)
						if m:
							indexRE=m.start()+indiceTourDeParole
							resteCalcule=texte[m.end()+1:]
						else:
							indexRE=-1
							
					#remplir la sous-requête (catégorisation des choses trouvées par find())
						sousRequetes= SR()
						for sr in sousRequetes:
							deb=indiceTourDeParole
							if indiceTourDeParole>7:
								deb=indiceTourDeParole-7								
							if texte[deb:indiceTourDeParole+10].find(sr)>0:
								laSousRequete=sr
					# mettre à jour reste et refaire un find() avant le prochain tour de boucle	
						reste=reste[prochaineOccurrenceDansReste+len(maRequete):]						
						prochaineOccurrenceDansReste=reste.find(maRequete)
						nbOcc=nbOcc+1
						resuCSV=resuCSV+"O;"+nomEnquete+";"+maRequete+";"+xDeb+";"+nomTire+";"+str(indiceTourDeParole)+";"+laSousRequete+";"+str(indexRE)+";"+texte+"\n"
					
					
				else:# c-à-d. si maRequete n'a pas été trouvé dans l'enquête.
					nbMotsEnquete=nbMotsEnquete#+nbMotsTourDeParole
	
					
	#print(cptLignes," lignes traitées") #a decommenter
#	print "Nombre d'occurrences de \"",maRequete,"\" : ",nbOcc	
	return([resuCSV,nomEnquete,cptLignes,maRequete])			
	fo.close()

import os
import pandas
from pathlib import Path
import openpyxl

def wraptext(sheet):
    """
	Cette fonction permet d'ajouter la propriété de passer à la ligne à chaque cellule de la feuille entré en paramètre.
    ----------
    sheet : WorkBook
		Feuille d'un fichier excel
    """
    from openpyxl.styles import Alignment
    for rows in sheet.iter_rows(min_row=1, min_col=1):
        for cell in rows:
            cell.alignment = Alignment(wrap_text=True,vertical='center')#Wraptext=passsage à la ligne

def fitauto(sheet):
    """
	Cette fonction permet d'ajuster la taille des colonnes en fonction de la taille de cellules max afin d'avoir un contenu
	homogène.
    ----------
    sheet : WorkBook
		Feuille d'un fichier excel
    """
    from openpyxl.utils.cell import _get_column_letter
    column_widths = []
    for row in sheet.iter_rows():
        for i, cell in enumerate(row):#Pour chaque cellule de la feuille par ligne
            try:
                size = len(str(cell.value))
                column_widths[i] = max(column_widths[i], size)#Compare la valeur de la largeur de colonne à celle de ces cellules
                # print(column_widths)
                #print(str(cell.value) + " && " + str(column_widths[i]) + " index: " + str(i))
            except IndexError:
                column_widths.append(len(str(cell.value)))
    print(column_widths)
    if column_widths.__len__() ==9:
        column_widths[8]=100#Changement de la largeur de colonne qui sera fixe pour la colonne tour de parole
    for i, column_width in enumerate(column_widths):
        # print(column_width)
        sheet.column_dimensions[_get_column_letter(
            i + 1)].width = column_width + 1#On fixe la taille max selon le tableau column_widths

def formatxlsx(path):
    """
	Cette fonction permet de bien formatter un fichier xlsx pour qu'elle soit présentable. Elle va appeler les fonctions
	wraptext et fitauto pour chaque feuille du fichier excel. A la fin il va sauvegarder les modifications dans ce même path
    Parameters
    ----------
    path : str
        Path du xlsx que l'on veut bien formatter
    """
    wb = openpyxl.load_workbook(path)
    wb["Data"].delete_cols(1)#Delete column of DataFrame Numbers are useless
    for name in wb.sheetnames:
        wraptext(wb[str(name)])
        fitauto(wb[str(name)])
    wb.save(path)	

def csvtoxlsx(path):
    """
	Cette fonction permet de convertir un csv en .xlsx. Elle permet en amont d'ordonner les données en classant la colonne ['présent']
	Mais aussi de créer un feuille où des statistiques sont rangés selon les tires/enquêtes et requêtes/sous-requêtes (permet d'avoir le nombre
	de requêtes/sous-requêtes en fonction des tires et de leurs enquêtes)
    Parameters
    ----------
    path : str
        Path du csv que l'on veut convertir en xlsx et ajouter la feuille de statistique
    """
    data_frame = pandas.read_csv(path, delimiter=";")
    data_frame = data_frame.sort_values(['présent'],ascending=False)#Ordonne les données en fonction de la colonne présent
    piv = data_frame.pivot_table(values=['requête', 'sous-requête'], index=['enquête', 'tire'], aggfunc={'requête':'count','sous-requête':'count'}, fill_value=0)
    new_filename = os.path.splitext(path)[0]+".xlsx"#Modification de l'extension du fichier
    with pandas.ExcelWriter(new_filename) as writer:  # doctest: +SKIP
    	data_frame.to_excel(writer, sheet_name='Data')
    	piv.to_excel(writer, sheet_name='DataResults')
    os.remove(path)#Supprimer le fichier .csv à la fin du traitement



##################################################################################################################################################################################################################################################################################### 

#PROGRAMME PRINCIPAL

##############################################################################################################################################################################################################################################################################
import os
import os.path
import sys
import expansions
import re
import datetime
import csv


###
###  PARAMETRES A METTRE A JOURï
###
print("*********************\n DEBUT EXECUTION \n*********************\n")
print("  PARAMèTRES\n")




# DATE
date=datetime.datetime.now().strftime("%y_%m_%d")
print("date (pour les résultats) = ",date,"\n")

#Repertoire où sont tous les fichiers à traiter
repertoireFichierTraite="corpus" # A DECOMMENTER EN CAS NORMAL

print("repertoireFichierTraite = ",repertoireFichierTraite,"\n")

# Repertoire où les fichiers résultats de la recherche sont mis 
nomRepertoireFichiersResultats= "./resultats/"+date
print("nomRepertoireFichiersResultats = ",nomRepertoireFichiersResultats,"\n")
nomRepertoireACreer = nomRepertoireFichiersResultats
print("Répertoire créé s'il n'existait pas :",nomRepertoireACreer,"\n")
if os.path.exists(nomRepertoireACreer):
	print("Le répertoire ",nomRepertoireACreer, " existait déjà")
else:
	os.mkdir(nomRepertoireACreer)

#creation si necessaire du repertoireFichierResultats
nomRepertoireACreer=  "./resultats/"+date+"/listesFichiersTraites"
if os.path.exists(nomRepertoireACreer):
	print("Le répertoire ",nomRepertoireACreer, " existait déjà")
else:
	os.mkdir(nomRepertoireACreer)

variableSearch = sys.argv[1]
fichierListe = date+"_"+"_"+variableSearch+"_listeEnquetes.csv"
print("fichierListe = ",fichierListe,"\n")

print("    FIN PARAMETRES \n**************************\n")

####
### FIN DES PARAMETRES A METTRE A JOUR
####

## Expansion de la requête donnée en argument		(on part de la liste vide)
listeRequetes=list()

## Recours au module expansions.py, dans lequel est définie exhaustivement l'expansion de la chaîne donnée en argument de l'appel du présent programme
listeRequetes=expansions.expansion(variableSearch)


print('Requetes de cette extraction : ',listeRequetes)

#################################################################
# RE-CREATION DU FICHIER CSV QUI CONTIENT LA LISTE DES Textgrids
# sur lesquels porte la recherche
#
#### récupération de la liste des fichiers dans le répertoire qui contient les textgrids ; 
print("Récupération de la liste des fichiers du répertoire où il y a le corpus (",repertoireFichierTraite,")\n")
listeTranscriptionsATraiter = os.listdir(repertoireFichierTraite) 
print ("nombre de fichiers moulinés :",len(listeTranscriptionsATraiter))

#
####Initialisation de la liste des fichiers récupérée plus haut dans un fichier listeEnquetes.csv
print("Initialisation du fichier listesFichiersTraites/",fichierListe,"\n")
listeFichiers = open("resultats/"+date+"/"+"listesFichiersTraites/"+fichierListe, 'w')
listeFichiers.write("Nom Enquête;requête; Nb. mots traités;\n")



# initialisation du fichier résultat (pour toutes les enquêtes du corpus ; la fonction retourRequete fabrique aussi un csv pour chaque enquête)
fichierResu=open(nomRepertoireFichiersResultats+"/"+variableSearch+".csv",'w')
fichierResu.write("présent;enquête;requête;temps;tire;stringIndex;sous-requête;indexRE;tour de parole\n")


for row in listeTranscriptionsATraiter: # boucle sur les transcriptions
	if not row.find(".DS_Store")>-1: # si dans la liste des fichiers du répertoire, il y a ".DS_Store", on ignore.
		for rr in listeRequetes: # boucle sur les requêtes. Pour chaque couple (transcription, requête), on récupère quatre valeurs : 
		#				* les lignes qui correspondent aux occurrences de la requête dans la transcription (retourRequete[0])
		#				* le nom de l'enquête (retourRequete[1])
		#				* le nombre de mots traités (retourRequête[2])
		#				* la requête(retourRequête[3])

			
			ret = retourRequete(\
			repertoireFichierTraite,\
			rr,row,nomRepertoireFichiersResultats,fichierResu)
			if ret[0] != '': # ici, si retourRequete a retourné un résultat, on inscrit dans le fichier résultat. 
									#Sinon, on inscrit dans le fichier résultat une ligne qui indique que la recherche a été faite pour cette transcription.
				fichierResu.write(ret[0]),"\n"
				#print(ret[0]+"\n")
			else:
				fichierResu.write("N"+";"+row+";"+rr+"\n")
		# Ecriture du fichier des requêtes traitées et du nombre de mots pour chacune 
		#(ce qui permet de vérifier que les fichiers ont été lus par le programme))	
			listeFichiers.write(str(ret[1])+";"+str(ret[3])+";"+str(ret[2])+";"+"\n")
				
fichierResu.close()

#Appel des fonction à la fin de la création du fichier .csv
csvtoxlsx(nomRepertoireFichiersResultats+"/"+variableSearch+".csv")
formatxlsx(nomRepertoireFichiersResultats+"/"+variableSearch+".xlsx")
